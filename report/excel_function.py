from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

class orangehrm_test_cases:
   def __init__(self, excel_file_name, sheet_name):
       self.file = excel_file_name
       self.sheet = sheet_name


   # fetch the row count
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   # fetch the column count
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   # read data from excel file
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


   # write data into excel file
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)

   def capture_text (self, xpath):
       wait = WebDriverWait(self.driver, 10)
       
       element = wait.until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
       value = element.text
